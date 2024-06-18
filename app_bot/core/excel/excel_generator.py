import io
import datetime
from openpyxl import Workbook
from tortoise.fields.relational import BackwardFKRelation
from tortoise.expressions import Q
from core.database.models import Advertisement, User


async def create_excel(model):
    entries = await model.all()

    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    # get model headers
    headers = []
    for field in model._meta.fields_map.values():
        if type(field) != BackwardFKRelation:
            headers.append(field.model_field_name)
    sheet.append(headers)

    # add users data
    for entry in entries:
        row = []
        for field_name in headers:
            cell = getattr(entry, field_name)
            if type(cell) == datetime.datetime:
                cell: datetime.datetime = cell.replace(tzinfo=None)
            row.append(cell)

        sheet.append(row)

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory


async def create_excel_for_agency(advertisements: list[Advertisement]):
    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    # create excel with data
    sheet.append(['ID', 'ТЗ', 'Цена', 'Прибыль менеджера', 'Чистая прибыль', 'Дата создания'])

    full_income, managers_income, clear_income = 0, 0, 0
    for adv in advertisements:
        manager: User = await adv.manager
        manager_income = round(adv.price * manager.manager_percent / 100)

        full_income += adv.price
        managers_income += manager_income

        sheet.append(
            [
                adv.id,
                adv.text,
                adv.price,
                manager_income,
                adv.price - manager_income,
                adv.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
        )

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory


async def create_excel_for_agency_with_managers_data(managers: list[User], agency_id: int):
    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    # create excel with data
    sheet.append(['ID', 'Username', 'Кол-во клиентов'])

    for manager in managers:
        adv_amount = await Advertisement.filter(Q(manager_id=manager.id) & ~Q(buyer_id=None) & Q(agency_id=agency_id)).count()

        sheet.append(
            [
                manager.id,
                manager.username,
                adv_amount,
            ]
        )

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory
